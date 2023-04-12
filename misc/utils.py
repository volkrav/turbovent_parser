from datetime import datetime
from typing import Dict, Generator
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import openpyxl
import csv
import datetime
import logging
import os
import sys
import aiofiles
import asyncio
import aiosqlite
import random
import pandas as pd

import aiohttp
import async_timeout
import pytz
from fake_useragent import UserAgent
from misc.classes import Too_Many_Requests

import settings


logger = logging.getLogger(__name__)
ua = UserAgent()


#! not used
# async def is_parse_categories(argv: list) -> bool:
#     if len(argv) > 1:
#         return argv[1] == 'parsecat'
with open('misc/workingproxies.txt', 'r') as file:
    print('\t worked open file')
    proxies_list = file.read().split()


async def fetch(session: aiohttp.ClientSession, url: str) -> None:
    headers = settings.headers
    headers['user-agent'] = ua.random
    # random.shuffle(proxies_list)
    for proxy in proxies_list:
        print(f'used {proxy} for {url}')
        try:
            proxies_list.append(proxies_list.pop(proxies_list.index(proxy)))
        except ValueError:
            logger.warning(f'{proxy} was moved earlier')
            continue
        try:
            async with async_timeout.timeout(settings.FETCH_TIMEOUT):
                async with session.get(url, headers=headers, proxy=proxy) as response:
                    if response.status == 200:
                        # print('get 200')
                        # sys.exit(1)
                        return await response.text()
                    elif response.status == 429:
                        logger.warning(
                            f'{url} get Too Many Requests, retrying in {settings.SLEEP_TIME}s')
                        await asyncio.sleep(settings.SLEEP_TIME)
                    else:
                        raise aiohttp.ClientResponseError(
                            response.request_info,
                            response.history,
                            status=response.status,
                            message=response.reason,
                            headers=response.headers
                        )
        except aiohttp.ClientHttpProxyError as err:
            logger.warning(
                f'{url} request failed with proxy {proxy} : {err}, will try another one')
            if proxies_list and proxy in proxies_list:
                proxies_list.remove(proxy)
            continue
        except aiohttp.ClientResponseError as err:
            logger.warning(
                f'{url} request failed with proxy {proxy} : {err}, will try another one')
            if proxies_list and proxy in proxies_list:
                proxies_list.remove(proxy)
            continue
        except aiohttp.ServerDisconnectedError as err:
            logger.warning(
                f'{url} request failed with proxy {proxy} : {err}, will try another one')
            if proxies_list and proxy in proxies_list:
                proxies_list.remove(proxy)
            continue
        except OSError as err:
            logger.warning(
                f'{url} request failed with proxy {proxy} : {err}, will try another one')
            # if proxies_list and proxy in proxies_list: proxies_list.remove(proxy)
            continue

        except Exception as err:
            logger.error(
                f'{url} request failed with proxy {proxy} : {err}', exc_info=True)
            sys.exit(1)
    logger.error(f'{url} request failed with all proxy servers')
    sys.exit(1)


async def _get_current_date() -> str:
    tz = pytz.timezone('Europe/Kiev')
    return datetime.datetime.now(tz).strftime('%Y-%m-%d')


async def write_to_csv(data: dict) -> None:
    current_date = await _get_current_date()
    try:
        file_exists = os.path.exists(f'outbox/change_{current_date}.csv')

        async with aiofiles.open(f'outbox/change_{current_date}.csv',
                                 mode='a' if file_exists else 'w',
                                 encoding='utf-8'
                                 ) as file:
            fieldnames = data.keys()
            writer = csv.DictWriter(file, fieldnames=fieldnames)

            if not file_exists:
                writer.writeheader()

            await writer.writerow(data)

    except Exception as e:
        logger.error(f"An error occurred: {e}")


async def write_to_excel(data: dict):
    # print(data)
    today = await _get_current_date()
    path = f'outbox/change_{today}.xlsx'
    if os.path.exists(path):
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        row_num = worksheet.max_row + 1
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        headers = list(data.keys())
        for col_num, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_num).value = header
        row_num = 2

    values = list(data.values())
    for col_num, value in enumerate(values, start=1):
        worksheet.cell(row=row_num, column=col_num).value = value

    workbook.save(path)


async def read_excel_file(file_path) -> Generator:
    '''
    Этот код прочитает excel файл и создаст генератор,
    который будет возвращать по одной строке из вашего
    файла при каждом вызове. Ключи этого словаря – заголовки столбцов.
    Они будут иметь тип str. Значения будут также иметь тип str.
    '''
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # получение заголовков столбцов
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        columns = [str(col) for col in row]

    # генерация строк
    async def get_row():
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = [str(col) for col in row]
            yield dict(zip(columns, values))
    return get_row

# row_gen = await read_excel_file('path/to/your/excel/file.xlsx')

# async for row in row_gen():
#     print(row)


async def make_csv_from_db_suppliers(db: aiosqlite.Connection):
    async with aiosqlite.connect('parser.db') as db:
        db.row_factory = aiosqlite.Row
        async with db.execute('select id_website, sku, title, price, available from suppliers') as cursor:
            count = 1
            async for row in cursor:
                # await write_to_csv(
                #     {
                #         'id_website': row['id_website'],
                #         'sku': row['sku'],
                #         'title': row['title'],
                #         'price': row['price'],
                #         'available': row['available']
                #     },
                # )
                await write_to_excel(
                    {
                        'id_website': row['id_website'],
                        'sku': row['sku'],
                        'title': row['title'],
                        'price': row['price'],
                        'available': row['available']
                    },
                )
                # if count == 2:
                #     break
                # count += 1


async def process_excel(filename: str) -> Dict:
    # список заголовков, которые нужно обработать
    columns_to_include = ['Column1', 'Column2', 'Column5']
    # чтение данных из файла excel
    df = pd.read_excel(filename)

    # создание списка словарей с нужными данными
    for _, row in df.iterrows():
        entry = {}
        for column in columns_to_include:
            entry[column] = row[column]
        yield entry


async def main():
    # async with aiosqlite.connect('parser.db') as db:
    # await make_csv_from_db_suppliers(db)
    today = await _get_current_date()
    # path = f'/Users/volodymyr/Downloads/da99f56c3add563b564de15dd3bf9f70.xlsx'
    path = f'/Users/volodymyr/Projects/turbovent_parser/inbox/67dd3a5b92c24f5aabdfcdc5bb327465.xlsx'
    row_gen = await read_excel_file(path)
    async for row in row_gen():
        product = {row['Артикул']}
        print(row['Артикул'])
        print(row['Артикул'])
        break


if __name__ == '__main__':
    loop = asyncio.get_event_loop()
    a1 = loop.create_task(main())
    loop.run_until_complete(a1)
